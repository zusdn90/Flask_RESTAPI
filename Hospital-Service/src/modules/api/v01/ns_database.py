import pymysql
import schedule
import openpyxl
import random
import string
import os

from openpyxl import Workbook, load_workbook
from faker import Faker
from flask import request
from flask_restplus import Namespace, Resource, fields, reqparse
from modules.setting import dbConnector
from modules.setting import log as logger

parser = reqparse.RequestParser()
api = Namespace('datas', description = "DATABASE API")

path = os.getcwd() + "/modules/data/"
filename = 'generate_data.xlsx'

""" 
#################################################################
DATABASE API 
#################################################################
""" 

class Result(object):
    returncode = 1
    stdout = ""
    stderr = ""

    def __init__(self, returncode, stdout, stderr):
        self.returncode = returncode
        self.stdout = stdout
        self.stderr = stderr

@api.route('/create')
class CreateData(Resource):
    def post(self):
        faker = Faker()

        write_wb = Workbook()
    
        write_ws = write_wb.active
        write_ws['A1'] = 'patient_code'
        write_ws['B1'] = 'patient_name'
        write_ws['C1'] = 'vital_code'
        write_ws['D1'] = 'vital_value'
        write_ws['E1'] = 'alert_info'
        write_ws['F1'] = 'admission_date'

        vital_code = ['v01','v02','v03','v04','v05']

        for i in range(100):
            write_ws.append([str(i+1), faker.name(),'v01',faker.random_int(),'',faker.date_this_year()])
        
        write_wb.save(path + filename)

        return {"status" : "200", "message": "Data create success..."}

@api.route('/insert')
class InsertData(Resource):
    def post(self):
        db_class = dbConnector.Database()

        try:
            sql = 'insert into Patient values(%s, %s, %s, %s, %s, %s)'
    
            wb = load_workbook(path + filename, data_only=True)
            ws = wb['Sheet']
    
            iter_rows = iter(ws.rows)                
            next(iter_rows)
                
            for row in iter_rows:
                db_class.execute(sql, (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value))

            wb.close()
            db_class.commit()

            return {"status" : "200", "message": "Data insert success..."}

        except Exception as e:            
            logger.error_log(str(e))
        finally:
            db_class.close()
            wb.close()


        

